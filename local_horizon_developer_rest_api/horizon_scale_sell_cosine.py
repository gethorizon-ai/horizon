
# %% [markdown]
# **Required Packages**

# %%
# %pip install openai
# %pip install bert-score
# %pip install sklearn
# %pip install openpyxl
# %pip install -U scikit-learn
# %pip install -U scikit-learn scipy matplotlib
# %pip install plotly
# %pip install python-dotenv
# %pip install matplotlib
# %pip install langchain --upgrade
# %pip install rouge
# %pip install faiss-cpu

# %%
from rouge import Rouge
from langchain import HuggingFaceHub
from langchain.llms import OpenAI
from langchain.prompts import FewShotPromptTemplate
from langchain.vectorstores import FAISS
from langchain.prompts.example_selector import MaxMarginalRelevanceExampleSelector
from langchain.embeddings import OpenAIEmbeddings
from langchain.schema import HumanMessage, AIMessage
from langchain.prompts.chat import HumanMessagePromptTemplate
from langchain.chat_models import ChatOpenAI
from langchain.prompts import PromptTemplate
from openpyxl import load_workbook
import json
import csv
import random
from dotenv import load_dotenv
import statistics
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import openpyxl
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import TfidfVectorizer
from openai.embeddings_utils import (
    get_embedding,
    cosine_similarity
)
import bert_score
import os
from sklearn.neighbors import NearestNeighbors
import matplotlib as plt
import numpy as np
import pandas as pd
import openai
import time
import copy
import re
from tabulate import tabulate
import tiktoken
from functools import partial
from itertools import repeat
import multiprocessing
from multiprocessing import Pool, freeze_support
from sklearn.cluster import KMeans
from scipy.spatial.distance import euclidean


# %%
load_dotenv()
openai.api_key = os.getenv('OPENAI_API_KEY')
HUGGINGFACEHUB_API_TOKEN = os.getenv('HUGGINGFACEHUB_API_TOKEN')

# %% [markdown]
# **Prompt Generation**
# %%


def prompt_generation_user_objective(experiment: dict, model_object: tuple, num_prompts: int, global_prompt_id: list) -> pd.DataFrame:
    """
    Generate prompt candidates based on user-provided objective and input variables
    """
    metaprompt = PromptTemplate(
        template="""You are an intelligent English professor. You will craft an instruction that I can give to my friend to accomplish the following objective:

OBJECTIVE: {objective}

I need to include the following input variables in the instruction:
INPUT VARIABLES: {input_variables}

What instruction should I give to my friend to accomplish the objective? Do not do the work for my friend, but rather craft an instruction so he can do it. Make sure the objective is clearly communicated in the instruction, along with any supporting information to help my friend do the best possible job. ONLY GIVE ME THE INSTRUCTION, NOTHING ELSE!!!

INSTRUCTION:""",
        input_variables=['objective', 'input_variables'])
    formatted_metaprompt = metaprompt.format(objective=experiment['user_objective'],
                                             input_variables=', '.join('{input_variable}'.format(input_variable='<' + input_var + '>') for input_var in experiment['input_variables']))

    prompt_suffix = """\n\n==\nBEGIN:\n\n"""
    for input_var in experiment['input_variables']:
        prompt_suffix += '<' + input_var + '>: {' + input_var + '}\n'
    prompt_suffix += '<OUTPUT>:'

    # metaprompt_model = ChatOpenAI(max_tokens=1000, temperature=0.7, n=num_prompts)
    metaprompt_model = OpenAI(model_name="text-davinci-003", temperature=0.4,
                              max_tokens=1000, n=num_prompts, best_of=num_prompts)

    if type(metaprompt_model) == ChatOpenAI:
        formatted_metaprompt = [HumanMessage(content=formatted_metaprompt)]
    responses = metaprompt_model.generate(
        [formatted_metaprompt]).generations[0]

    prompt_id_list = []
    generation_id_list = []
    prompt_object_list = []
    prompt_prefix_list = []
    model_object_list = []
    for i in range(num_prompts):
        # Check that prompt template has required input variables and is formatted correctly
        prompt_prefix = responses[i].text.strip()
        prompt_template = prompt_prefix + prompt_suffix
        try:
            generated_prompt = PromptTemplate(
                template=prompt_template, input_variables=experiment['input_variables'])
        except:
            continue

        prompt_id_list.append(global_prompt_id[0])
        global_prompt_id[0] += 1
        generation_id_list.append('[user_objective]')
        prompt_object_list.append(tuple([generated_prompt]))
        prompt_prefix_list.append(prompt_prefix)
        model_object_list.append(copy.deepcopy(model_object))

    result = pd.DataFrame({
        'prompt_id': prompt_id_list,
        'generation_id': generation_id_list,
        'prompt_object': prompt_object_list,
        'prompt_prefix': prompt_prefix_list,
        'model_object': model_object_list
    })
    return result

# %%

def prompt_generation_user_objective_training_data(experiment: dict, model_object: tuple, num_prompts: int, global_prompt_id: list) -> pd.DataFrame:
    """
    Generate prompt candidates based on user-provided objective, input variables, and training data
    """
    # prefix to few shot-based metaprompt
    few_shot_metaprompt_prefix = """You are an intelligent English professor. You will craft an instruction that I can give to my friend to accomplish a task. Here is my task objective, input variables to be used, and ideal example outputs.

[BEGIN DATA]
==
OBJECTIVE: {objective}
INPUT VARIABLES: {input_variables}

==
EXAMPLES:

"""

    # suffix
    few_shot_metaprompt_suffix = """==
[END DATA]

What is the optimal instruction I should give to my friend to best accomplish my objective and generate output like in the examples? Do not do the work for my friend, but rather craft an instruction so she can do it. Do not refer to the examples above. Make sure the objective is clearly communicated in the instruction, along with any supporting information to help my friend do the best possible job. ONLY GIVE ME THE INSTRUCTION, NOTHING ELSE!!!

INSTRUCTION:"""

    #create a list of few shot examples. Each example should be a dictionary with the keys being the input variables and the values being the values for those input variables
    examples = cluster_shortlist_data(experiment, 5, 'train')

    #iterate through each input variable to create example prompt template
    exampleFormatterTemplate = ""
    for j in range (len(experiment['input_variables'])):
        exampleFormatterTemplate += "<" + experiment['input_variables'][j] + ">: {" + experiment['input_variables'][j] + "}\n"
    exampleFormatterTemplate += "<OUTPUT>: {output}"
    example_input_variables = experiment['input_variables'] + ['output']
    example_prompt = PromptTemplate(
        input_variables=example_input_variables,
        template=exampleFormatterTemplate
    )

    #create the few shot prompt template
    few_shot_metaprompt = FewShotPromptTemplate(
        examples=examples,
        example_prompt=example_prompt,
        prefix=few_shot_metaprompt_prefix,
        suffix=few_shot_metaprompt_suffix,
        input_variables=['objective', 'input_variables'],
    )

    formatted_metaprompt = few_shot_metaprompt.format(objective=experiment['user_objective'],
                                                      input_variables=', '.join('{input_variable}'.format(input_variable='<' + input_var + '>') for input_var in experiment['input_variables']))

    # metaprompt_model = ChatOpenAI(max_tokens=1000, temperature=0.7, n=num_prompts)
    metaprompt_model = OpenAI(model_name="text-davinci-003", temperature=0.4,
                              max_tokens=1000, n=num_prompts, best_of=num_prompts)

    if type(metaprompt_model) == ChatOpenAI:
        formatted_metaprompt = [HumanMessage(content=formatted_metaprompt)]
    responses = metaprompt_model.generate([formatted_metaprompt]).generations[0]

    prompt_id_list = []
    generation_id_list = []
    prompt_object_list = []
    prompt_prefix_list = []
    model_object_list = []

    prompt_suffix = """\n\n==\nBEGIN:\n\n"""
    for input_var in experiment['input_variables']:
        prompt_suffix += '<' + input_var + '>: {' + input_var + '}\n'
    prompt_suffix += '<OUTPUT>:'

    for i in range(num_prompts):
        # Check that prompt template has required input variables and is formatted correctly
        prompt_prefix = responses[i].text.strip()
        prompt_template = prompt_prefix + prompt_suffix
        try:
            generated_prompt = PromptTemplate(
                template=prompt_template, input_variables=experiment['input_variables'])
        except:
            continue

        prompt_id_list.append(global_prompt_id[0])
        global_prompt_id[0] += 1
        generation_id_list.append('[user_objective_training_data]')
        prompt_object_list.append(tuple([generated_prompt]))
        prompt_prefix_list.append(prompt_prefix)
        model_object_list.append(copy.deepcopy(model_object))

    result = pd.DataFrame({
        'prompt_id': prompt_id_list,
        'generation_id': generation_id_list,
        'prompt_object': prompt_object_list,
        'prompt_prefix': prompt_prefix_list,
        'model_object': model_object_list
    })
    return result

# %%


def prompt_generation_pattern_role_play(experiment: pd.DataFrame, model_object: tuple, num_prompts: int, global_prompt_id: list) -> pd.DataFrame:
    """
    Generates prompt candidates by converting provided prompt candidates into a role play prompt pattern
    """
    metaprompt = PromptTemplate(template="""You are a prompt generation robot that generates optimal prompt template strings for use with AI large language models (LLM). One effective approach for prompts is to frame them as a role play for the LLM. Therefore, you will provide the optimal role play prompt template string to accomplish the given objective using the given input variables. Surround input variables by angle brackets when referencing them.
==
EXAMPLES:

OBJECTIVE: Answer the philosophy question
INPUT VARIABLES: <question>
OPTIMAL PROMPT: I want you to act as a philosopher. I will provide some topics or questions related to the study of philosophy, and it will be your job to explore these concepts in depth. This could involve conducting research into various philosophical theories, proposing new ideas or finding creative solutions for solving complex problems. My first request is <question>

OBJECTIVE: Help me stay motivated given my situation
INPUT VARIABLES: <situation>
OPTIMAL PROMPT: I want you to act as a self-help book. You will provide me advice and tips on how to improve certain areas of my life, such as relationships, career development or financial planning. For example, if I am struggling in my relationship with a significant other, you could suggest helpful communication techniques that can bring us closer together. My current issue is {{situation}}

OBJECTIVE: Respond to my sentence as would the given character from the given series
INPUT VARIABLES: <character>, <series>, <sentence>
OPTIMAL PROMPT: I want you to act like <character> from <series>. I want you to respond and answer like <character> using the tone, manner and vocabulary <character> would use. Do not write any explanations. Only answer like <character>. You must know all of the knowledge of <character>. My first sentence is <sentence>

OBJECTIVE: Provide an advertising strategy for my request
INPUT VARIABLES: <request>
OPTIMAL PROMPT: I want you to act as an advertiser. You will create a campaign to promote a product or service of your choice. You will choose a target audience, develop key messages and slogans, select the media channels for promotion, and decide on any additional activities needed to reach your goals. My first suggestion request is <request>

==
BEGIN:

OBJECTIVE: {objective}
INPUT VARIABLES: {input_variables}
OPTIMAL PROMPT:""",
                                input_variables=['objective', 'input_variables'])
    formatted_metaprompt = metaprompt.format(objective=experiment['user_objective'],
                                             input_variables=', '.join('{input_variable}'.format(input_variable='<' + input_var + '>') for input_var in experiment['input_variables']))

    prompt_suffix = """\n\n==\nBEGIN:\n\n"""
    for input_var in experiment['input_variables']:
        prompt_suffix += '<' + input_var + '>: {' + input_var + '}\n'
    prompt_suffix += '<OUTPUT>:'

    # metaprompt_model = ChatOpenAI(max_tokens=1000, temperature=0.7, n=num_prompts)
    metaprompt_model = OpenAI(model_name="text-davinci-003", temperature=0.4,
                              max_tokens=1000, n=num_prompts, best_of=num_prompts)

    if type(metaprompt_model) == ChatOpenAI:
        formatted_metaprompt = [HumanMessage(content=formatted_metaprompt)]
    responses = metaprompt_model.generate(
        [formatted_metaprompt]).generations[0]

    prompt_id_list = []
    generation_id_list = []
    prompt_object_list = []
    prompt_prefix_list = []
    model_object_list = []
    for i in range(num_prompts):
        # Check that prompt template has required input variables and is formatted correctly
        prompt_prefix = responses[i].text.strip()
        prompt_template = prompt_prefix + prompt_suffix
        try:
            generated_prompt = PromptTemplate(
                template=prompt_template, input_variables=experiment['input_variables'])
        except:
            continue

        prompt_id_list.append(global_prompt_id[0])
        global_prompt_id[0] += 1
        prompt_prefix_list.append(prompt_prefix)
        generation_id_list.append('[pattern_role_play]')
        prompt_object_list.append(tuple([generated_prompt]))
        model_object_list.append(copy.deepcopy(model_object))

    result = pd.DataFrame({
        'prompt_id': prompt_id_list,
        'generation_id': generation_id_list,
        'prompt_object': prompt_object_list,
        'prompt_prefix': prompt_prefix_list,
        'model_object': model_object_list
    })
    return result

# %%
# create the example dataset for FewShotPromptTemplate


def get_example_dataset(experiment: pd.DataFrame) -> list:
    examples = []
    for i in range(len(experiment['ground_truth_few_shots'])):
        example = {}
        for j in range(len(experiment['input_variables'])):
            example[experiment['input_variables'][j]
                    ] = experiment['input_values_few_shots'][j].iloc[i]
        example['output'] = experiment['ground_truth_few_shots'].iloc[i]
        examples.append(example)
    return examples

# %%


def few_shots(experiment: pd.DataFrame, prompt_candidates: pd.DataFrame, num_few_shots: int, global_prompt_id: list) -> pd.DataFrame:

    prompt_id_list = []
    generation_id_list = []
    prompt_object_list = []
    prompt_prefix_list = []
    model_object_list = []

    # create a list of few shot examples. Each example should be a dictionary with the keys being the input variables and the values being the values for those input variables
    examples = get_example_dataset(experiment)

    # iterate through each - candidate
    for i in range(len(prompt_candidates)):
        few_shot_prompt_prefix = prompt_candidates['prompt_prefix'][i] + "\n\n==\nEXAMPLES:"
        # iterate through each input variable
        exampleFormatterTemplate = ""
        for j in range(len(experiment['input_variables'])):
            exampleFormatterTemplate += "<" + \
                experiment['input_variables'][j] + \
                ">: {" + experiment['input_variables'][j] + "}\n"

        suffix_request = "==\nBEGIN:\n\n" + exampleFormatterTemplate + "<OUTPUT>:"
        # add the output variable
        exampleFormatterTemplate += "<OUTPUT>: {output}"
        new_input_variables = experiment['input_variables'] + ['output']
        # create the prompt template
        example_prompt = PromptTemplate(
            input_variables=new_input_variables,
            template=exampleFormatterTemplate,
        )

        # create the example selector
        # examples: This is the list of examples available to select from.
        # embeddings: This is the Embeddings class that is used to embed the examples.
        # vector_store: This is the VectorStore class that is used to store the embeddings and do a similarity search over.
        # k: This is the number of examples to produce.
        example_selector = MaxMarginalRelevanceExampleSelector.from_examples(
            examples, OpenAIEmbeddings(), FAISS, k=num_few_shots)

        # create the few shot prompt template
        few_shot_prompt = FewShotPromptTemplate(
            example_selector=example_selector,
            example_prompt=example_prompt,
            prefix=few_shot_prompt_prefix,
            suffix=suffix_request,
            input_variables=experiment['input_variables'],
        )

        # add the few shot prompt to the prompt_candidates list
        prompt_id_list.append(global_prompt_id[0])
        global_prompt_id[0] += 1
        generation_id_list.append(
            prompt_candidates['generation_id'][i] + '_[few_shots_max_marginal_relevance]')
        prompt_object_list.append(tuple([few_shot_prompt]))
        prompt_prefix_list.append(prompt_candidates['prompt_prefix'][i])
        model_object_list.append(copy.deepcopy(
            prompt_candidates['model_object'][i]))

    result = pd.DataFrame({
        'prompt_id': prompt_id_list,
        'generation_id': generation_id_list,
        'prompt_object': prompt_object_list,
        'prompt_prefix': prompt_prefix_list,
        'model_object': model_object_list
    })
    return result

# %%

def prompt_generation_targeted_improvement(experiment: pd.DataFrame, prompt_candidates: pd.DataFrame, global_prompt_id: list) -> pd.DataFrame:
    """
    Generate new prompt for each of the provided prompts that corrects errors. Only returns the new prompts.
    """
    unique_prompt_ids = prompt_candidates['prompt_id'].unique().tolist()
    inference_results = run_inference(experiment, prompt_candidates, 'train')
    evaluation_results = run_evaluation(inference_results)

    metaprompt_assessment = """You are an analyzing bot. I used the following instruction string template with a friend.

INSTRUCTION STRING: {original_prompt}

INPUT VARIABLES: Here are the input values that I substituted into the instruction string template before giving it to my friend:\n\n"""
    for input_var in experiment['input_variables']:
        metaprompt_assessment += '<' + input_var + '>: ' + '{' + input_var + '}\n'
    metaprompt_assessment += """\nYOUR TASK: The friend read the instruction string template and wrote an output. Here is the actual output from the friend and the ideal output that I desired. Conduct a thorough and detailed evaluation of the differences between the actual output and ideal output. Consider all conceptual, structural, and formatting differences. Work it out in a step by step way to be sure you have the right answer. Then, provide a detailed and actionable recommendation for me to improve the instruction string template. Do not suggest changes that are specific to these examples. Do not suggest providing any additional examples or changes to the input variables. Do not provide an example of a new instruction string.

ACTUAL OUTPUT: {output}
IDEAL OUTPUT: {ground_truth}

ANALYSIS AND RECOMMENDATION:"""

    metaprompt_synthesis = """You are an insightful and award-winning English professor. I created an instruction for my friend and shared it with my peers for feedback. My peers shared the following feedback on what worked well and how I should improve my instruction. Provide an actionable, insightful, and concise synthesis of their feedback that I can use to improve my instruction. Include feedback that is consistent across all of them while avoiding feedback that is narrow or only specific to one of my peers.

FEEDBACK 1: {feedback1}

FEEDBACK 2: {feedback2}

FEEDBACK 3: {feedback3}

SYNTHESIS:"""

    metaprompt_update = """You are an insightful and award-winning English professor. I used the following original instruction string with a friend. My friends then gave me feedback on how I can improve the instruction. Craft a better instruction that incorporates their recommendation below. Keep the context and relevant details from my original instruction.

ORIGINAL INSTRUCTION: {original_prompt}

RECOMMENDATION: {recommendation}

BETTER INSTRUCTION:"""

    metaprompt_check = """You are an intelligent English professor. I read aloud a very sensitive instruction for my colleague. My friend heard it and tried to write down the instruction. Here is my original instruction and my friend's version:

[BEGIN DATA]
==
ORIGINAL INSTRUCTION:
{original_prompt}

==
FRIEND'S VERSION:
{new_prompt}

==
[END DATA]

What new information did my friend's version have that was not previously provided? Check if any new facts or information is communicated beyond clarifications or improvements in style or formatting. Work it out in a detailed step by step manner to be sure you have the right answer. Think step-by-step, then provide your explanations and final answer as a JSON instance that conforms to the following JSON schema:

{{"observations": {{"description": "detailed observations of new information included in my friend's version"}}, "type": "string"}},
"analysis": {{"description": "is the new information a clarification of my original instructions, such as formatting or stylistic suggestions, or does it contain new data or facts that may not be true?", "type": "string"}}
"final_answer": {{"description": "based on the analysis above, final answer as 'YES' or 'NO' if my friend's version included any new information beyond clarifications", "type": "string", "enum": ["YES", "NO"]}}}}

JSON OUTPUT:"""

    metaprompt_model_generation = OpenAI(
        model_name="text-davinci-003", temperature=0.7, max_tokens=500)
    metaprompt_model_check = OpenAI(
        model_name="text-davinci-003", temperature=0.3, max_tokens=500)    

    metaprompt_assessment_template = PromptTemplate(template=metaprompt_assessment, input_variables=[
                                                    'original_prompt'] + experiment['input_variables'] + ['output', 'ground_truth'])
    metaprompt_synthesis_template = PromptTemplate(template=metaprompt_synthesis, input_variables=[
                                                   'feedback1', 'feedback2', 'feedback3'])
    metaprompt_update_template = PromptTemplate(
        template=metaprompt_update, input_variables=['original_prompt', 'recommendation'])
    metaprompt_check_template = PromptTemplate(
        template=metaprompt_check, input_variables=['original_prompt', 'new_prompt'])

    prompt_id_list = []
    generation_id_list = []
    prompt_object_list = []
    prompt_prefix_list = []
    model_object_list = []

    rows_to_evaluate = 3
    num_tries = 3

    for prompt_id in unique_prompt_ids:
        for run in range(num_tries):
            # Get inferences with worst score
            error_rows = evaluation_results[evaluation_results['prompt_id'] == prompt_id].nsmallest(
                rows_to_evaluate, 'cosine_similarity_openAI').reset_index(drop=True)

            # Extract prefix, prompt object, and model object of original prompt
            original_prompt_prefix = error_rows['prompt_prefix'][0]
            original_prompt_object = error_rows['prompt_object'][0][0]
            original_model_object = error_rows['model_object'][0][0]

            # Generate feedback on each error row
            feedback = {'feedback1': '', 'feedback2': '', 'feedback3': ''}
            for row_index in range(rows_to_evaluate):
                metaprompt_assessment_formatted = metaprompt_assessment_template.format(**({'original_prompt': original_prompt_prefix} |
                                                                                        error_rows['input_values'][row_index] |
                                                                                        {'output': error_rows['output'][row_index], 'ground_truth': error_rows['ground_truth'][row_index]}))

                response = metaprompt_model_generation.generate(
                    [metaprompt_assessment_formatted]).generations[0][0].text.strip()
                feedback['feedback' + str(row_index + 1)] = response

            # Synthesize feedback across error rows
            metaprompt_synthesis_formatted = metaprompt_synthesis_template.format(**feedback)
            synthesis = metaprompt_model_generation.generate(
                [metaprompt_synthesis_formatted]).generations[0][0].text.strip()

            # Generate new prompt prefix
            metaprompt_update_formatted = metaprompt_update_template.format(
                original_prompt=original_prompt_prefix, recommendation=synthesis)
            new_prompt_prefix = metaprompt_model_generation.generate(
                [metaprompt_update_formatted]).generations[0][0].text.strip()
            
            # Check if new generated prompt prefix is overfitting
            metaprompt_check_formatted = metaprompt_check_template.format(
                original_prompt=original_prompt_prefix, new_prompt=new_prompt_prefix)
            overfit_assessment = metaprompt_model_check.generate(
                [metaprompt_check_formatted]).generations[0][0].text.strip()
            try:
                overfit_check = json.loads(overfit_assessment)
                assert overfit_check['final_answer'] in ['YES', 'NO']
                if overfit_check['final_answer'] == 'YES':
                    continue
            except:
                continue

            # Check if original prompt object is FewShotPromptTemplate. If so, update accordingly
            if isinstance(original_prompt_object, FewShotPromptTemplate):
                # Extract original example selector, example prompt, and suffix
                original_example_selector = original_prompt_object.example_selector
                original_example_prompt = original_prompt_object.example_prompt
                original_suffix = original_prompt_object.suffix

                # Create new few shot prefix
                new_few_shot_prompt_prefix = new_prompt_prefix + "\n\n==\nEXAMPLES:\n\n"

                # Create new few shot prompt template
                try:
                    generated_prompt = FewShotPromptTemplate(
                        example_selector=original_example_selector,
                        example_prompt=original_example_prompt,
                        prefix=new_few_shot_prompt_prefix,
                        suffix=original_suffix,
                        input_variables=experiment['input_variables'],
                    )
                except:
                    continue

            # Check if original prompt object is PromptTemplate. If so, update accordingly
            elif isinstance(original_prompt_object, PromptTemplate):
                prompt_suffix = """\n\n==\nBEGIN:\n\n"""
                for input_var in experiment['input_variables']:
                    prompt_suffix += '<' + input_var + '>: {' + input_var + '}\n'
                prompt_suffix += '<OUTPUT>:'

                new_prompt_string = new_prompt_prefix + prompt_suffix
                try:
                    generated_prompt = PromptTemplate(
                        template=new_prompt_string, input_variables=experiment['input_variables'])
                except:
                    continue

            prompt_id_list.append(global_prompt_id[0])
            global_prompt_id[0] += 1
            generation_id_list.append(error_rows['generation_id'][0] + '_[TIM]')
            prompt_object_list.append(tuple([generated_prompt]))
            prompt_prefix_list.append(new_prompt_prefix)
            model_object_list.append(tuple([copy.deepcopy(original_model_object)]))
            break

    result = pd.DataFrame({
        'prompt_id': prompt_id_list,
        'generation_id': generation_id_list,
        'prompt_object': prompt_object_list,
        'prompt_prefix': prompt_prefix_list,
        'model_object': model_object_list
    })
    return result

# %%


def prompt_generation_temperature_variation(experiment: dict, prompt_candidates: pd.DataFrame, num_prompts: int, global_prompt_id: list) -> pd.DataFrame:
    """
    For each of the given prompt candidates, generate new prompt candidates with same prompt but varying model temperatures
    """
    temperature_trials = np.linspace(0.1, 0.9, num_prompts)

    prompt_id_list = []
    generation_id_list = []
    prompt_object_list = []
    prompt_prefix_list = []
    model_object_list = []

    for index, row in prompt_candidates.iterrows():
        for temperature in temperature_trials:
            prompt_id_list.append(global_prompt_id[0])
            global_prompt_id[0] += 1

            generation_id_list.append(row['generation_id'] + '_[temperature_variation]')

            prompt_object_list.append(copy.deepcopy(row['prompt_object']))
            prompt_prefix_list.append(copy.deepcopy(row['prompt_prefix']))

            selected_model = row['model_object'][0]
            new_model = copy.deepcopy(selected_model)
            if isinstance(selected_model, ChatOpenAI):
                new_model.model_kwargs['temperature'] = temperature
            elif isinstance(selected_model, OpenAI):
                new_model.temperature = temperature
            elif isinstance(selected_model, HuggingFaceHub):
                new_model.model_kwargs['temperature'] = temperature
            model_object_list.append(tuple([new_model]))

    result = pd.DataFrame({
        'prompt_id': prompt_id_list,
        'generation_id': generation_id_list,
        'prompt_object': prompt_object_list,
        'prompt_prefix': prompt_prefix_list,
        'model_object': model_object_list
    })
    return result

# %%


def prompt_generation_variants(experiment: dict, prompt_candidates: pd.DataFrame, num_prompts: int, global_prompt_id: list) -> pd.DataFrame:
    """
    Generates syntactic variants of the given prompts that are semantically similar to the original. Assumes prompts are PromptTemplate objects.
    """
    metaprompt = PromptTemplate(template="""You are a creative writer. I came up with an instruction for my friend to complete. Generate a creative variation of my original instruction while keeping the same exact semantic meaning and objective. Include all relevant context and details from my original instruction so my friend completes the specific intent of my original instruction. Do not complete the instruction or provide an example completion.

ORIGINAL INSTRUCTION: {prompt_template}
NEW INSTRUCTION:""",
        input_variables=['prompt_template'])
    
    metaprompt_check_template = PromptTemplate(template="""You are an intelligent English professor. I read aloud a very sensitive instruction for my colleague. My friend heard it and tried to write down the instruction. Here is my original instruction and my friend's version:

[BEGIN DATA]
==
ORIGINAL INSTRUCTION:
{original_prompt}

==
FRIEND'S VERSION:
{new_prompt}

==
[END DATA]

What new information did my friend's version have that was not previously provided? Check if any new facts or information is communicated beyond clarifications or improvements in style or formatting. Work it out in a detailed step by step manner to be sure you have the right answer. Think step-by-step, then provide your explanations and final answer as a JSON instance that conforms to the following JSON schema:

{{"observations": {{"description": "detailed observations of new information included in my friend's version"}}, "type": "string"}},
"analysis": {{"description": "is the new information a clarification of my original instructions, such as formatting or stylistic suggestions, or does it contain new data or facts that may not be true?", "type": "string"}}
"final_answer": {{"description": "based on the analysis above, final answer as 'YES' or 'NO' if my friend's version included any new information beyond clarifications", "type": "string", "enum": ["YES", "NO"]}}}}

JSON OUTPUT:""",
        input_variables=['original_prompt', 'new_prompt'])

    prompt_suffix = """\n\n==\nBEGIN:\n\n"""
    for input_var in experiment['input_variables']:
        prompt_suffix += '<' + input_var + '>: {' + input_var + '}\n'
    prompt_suffix += '<OUTPUT>:'

    # metaprompt_model = ChatOpenAI(max_tokens=1000, temperature=0.6, n=num_prompts)
    metaprompt_model_generation = OpenAI(model_name="text-davinci-003", temperature=0.9,
                              max_tokens=1000, n=num_prompts, best_of=num_prompts)
    
    metaprompt_model_check = OpenAI(model_name="text-davinci-003", temperature=0.3,
                              max_tokens=500)

    prompt_id_list = []
    generation_id_list = []
    prompt_object_list = []
    prompt_prefix_list = []
    model_object_list = []

    for index, row in prompt_candidates.iterrows():
        original_prompt_prefix = row['prompt_prefix']
        model_object = row['model_object']

        formatted_metaprompt = metaprompt.format(
            prompt_template=original_prompt_prefix)

        if type(metaprompt_model_generation) == ChatOpenAI:
            formatted_metaprompt = [HumanMessage(content=formatted_metaprompt)]
        responses = metaprompt_model_generation.generate([formatted_metaprompt]).generations[0]

        for i in range(num_prompts):
            new_prompt_prefix = responses[i].text.strip()
            prompt_template = new_prompt_prefix + prompt_suffix

            # Check that prompt template has required input variables and is formatted correctly
            try:
                generated_prompt = PromptTemplate(
                    template=prompt_template, input_variables=experiment['input_variables'])
            except:
                continue
            
            # Check if new generated prompt is overfitting
            metaprompt_check_formatted = metaprompt_check_template.format(
                original_prompt=original_prompt_prefix, new_prompt=new_prompt_prefix)
            overfit_assessment = metaprompt_model_check.generate(
                [metaprompt_check_formatted]).generations[0][0].text.strip()
            try:
                overfit_check = json.loads(overfit_assessment)
                assert overfit_check['final_answer'] in ['YES', 'NO']
                if overfit_check['final_answer'] == 'YES':
                    continue
            except:
                continue

            prompt_id_list.append(global_prompt_id[0])
            global_prompt_id[0] += 1
            generation_id_list.append(
                row['generation_id'] + '_[variant]')
            prompt_object_list.append(tuple([generated_prompt]))
            prompt_prefix_list.append(new_prompt_prefix)
            model_object_list.append(copy.deepcopy(model_object))

        result = pd.DataFrame({
            'prompt_id': prompt_id_list,
            'generation_id': generation_id_list,
            'prompt_object': prompt_object_list,
            'prompt_prefix': prompt_prefix_list,
            'model_object': model_object_list
        })
        return result

    # %%

def cluster_shortlist(prompt_candidates: pd.DataFrame, num_clusters: int) -> pd.DataFrame:
    """
    Return a shortlisted version of the given prompt_candidates by clustering prompt templates and picking the prompt closest to the centroid of each cluster
    """
    # Function to calculate embedding of each prompt template string
    def calculate_embedding(prompt_object: tuple) -> float:
        original_prompt = prompt_object[0]
        if isinstance(original_prompt, FewShotPromptTemplate):
            template_string = original_prompt.prefix
        elif isinstance(original_prompt, PromptTemplate):
            template_string = original_prompt.template
        return OpenAIEmbeddings().embed_query(template_string)

    # Compute k-means clusters of prompt templates
    prompt_template_embeddings = list(
        prompt_candidates['prompt_object'].apply(lambda x: calculate_embedding(x)))
    clusters = KMeans(n_clusters=num_clusters).fit(prompt_template_embeddings)

    # Loop over all clusters and find index of closest point to the cluster center and append to closest_prompt_index list
    closest_prompt_index = []
    for cluster_index in range(num_clusters):
        # Get all indices of points assigned to this cluster:
        cluster_points_indices = np.where(clusters.labels_ == cluster_index)[0]

        cluster_center = clusters.cluster_centers_[cluster_index]
        min_index = np.argmin([euclidean(prompt_template_embeddings[cluster_point_index],
                              cluster_center) for cluster_point_index in cluster_points_indices])
        closest_prompt_index.append(cluster_points_indices[min_index])

    # Shortlist prompt candidates to those closest to the centroid of each cluster
    shortlisted_prompt_candidates = prompt_candidates.filter(
        items=closest_prompt_index, axis=0).reset_index(drop=True)
    return shortlisted_prompt_candidates

# %%

def cluster_shortlist_data(experiment: dict, num_clusters: int, train_or_test: str) -> list:
    """
    Embeds input and output data and returns num_clusters of them as list of dicts
    """
    # Select input and ground truth data from train or test dataset
    if train_or_test == 'train':
        input_values = experiment['input_values_train']
        ground_truth = experiment['ground_truth_train']
    elif train_or_test == 'test':
        input_values = experiment['input_values_test']
        ground_truth = experiment['ground_truth_test']

    # Build list of dictionaries for example data
    examples = []
    for i in range(len(ground_truth)):
        example = {}
        for j in range(len(experiment['input_variables'])):
            example[experiment['input_variables'][j]
                    ] = input_values[j].iloc[i]
        example['output'] = ground_truth.iloc[i]
        examples.append(example)

    # Function to calculate embedding of each example
    def calculate_embedding(example: dict) -> float:
        example_string = ''
        for key, value in example.items():
            example_string += '<{key}>: {value}\n'.format(key=key, value=value)
        example_string = example_string.strip()
        return OpenAIEmbeddings().embed_query(example_string)

    # Embed examples and compute k-means clusters of them
    example_embeddings = [calculate_embedding(example) for example in examples]
    clusters = KMeans(n_clusters=num_clusters).fit(example_embeddings)

    # Loop over all clusters and find index of closest point to the cluster center and append to closest_prompt_index list
    closest_prompt_index = []
    for cluster_index in range(num_clusters):
        # Get all indices of points assigned to this cluster:
        cluster_points_indices = np.where(clusters.labels_ == cluster_index)[0]

        cluster_center = clusters.cluster_centers_[cluster_index]
        min_index = np.argmin([euclidean(example_embeddings[cluster_point_index],
                              cluster_center) for cluster_point_index in cluster_points_indices])
        closest_prompt_index.append(cluster_points_indices[min_index])

    # Shortlist examples to those closest to the centroid of each cluster
    shortlisted_examples = [examples[index] for index in closest_prompt_index]
    return shortlisted_examples

# %%

def run_single_inference(prompt_id: int, generation_id: int, prompt_object: tuple, prompt_prefix: str, model_object: tuple, input_values: dict, ground_truth: str) -> dict:
    """
    Completes a single inference with the given prompt template by substituting the input values for the input variable placeholders
    """
    input_values = copy.deepcopy(input_values)
    formatted_prompt = prompt_object[0].format(**input_values)
    model = model_object[0]

    print('Working on prompt ID: ' + str(prompt_id))

    if type(model) == ChatOpenAI:
        formatted_prompt = [HumanMessage(content=formatted_prompt)]

    start_time = time.time()
    output = model.generate([formatted_prompt]).generations[0][0].text.strip()
    end_time = time.time()
    latency = end_time - start_time

    result = {
        'prompt_id': prompt_id,
        'generation_id': generation_id,
        'prompt_object': prompt_object,
        'prompt_prefix': prompt_prefix,
        'model_object': model_object,
        'input_values': input_values,
        'output': output,
        'ground_truth': ground_truth,
        'latency': latency
    }
    return result

# %%


def run_inference(experiment: pd.DataFrame, prompt_candidates: pd.DataFrame, train_or_test_dataset: str) -> pd.DataFrame:
    """
    Run inference with given set of prompt candidates and either train or test dataset
    """
    inferences = []
    input_values_iter = dict((input_variable, '')
                             for input_variable in experiment['input_variables'])
    for index, prompt_candidate_iter in prompt_candidates.iterrows():
        if train_or_test_dataset == 'test':
            input_values = experiment['input_values_test']
            ground_truth = experiment['ground_truth_test']
        elif train_or_test_dataset == 'train':
            input_values = experiment['input_values_train']
            ground_truth = experiment['ground_truth_train']

        for input_value_index in range(len(input_values[0])):
            for input_variable_index in range(len(experiment['input_variables'])):
                input_values_iter[experiment['input_variables'][input_variable_index]
                                  ] = input_values[input_variable_index].iloc[input_value_index]
            result = run_single_inference(prompt_candidate_iter['prompt_id'], prompt_candidate_iter['generation_id'], prompt_candidate_iter['prompt_object'],
                                          prompt_candidate_iter['prompt_prefix'], prompt_candidate_iter['model_object'],
                                          input_values_iter, ground_truth.iloc[input_value_index])
            inferences.append(result)
    result = pd.DataFrame(inferences)
    return result

# %% [markdown]
# **Prompt Evaluation**

# %%


def get_semantic_cosine_similarity_openAI(df):
    scores = []
    for i, row in df.iterrows():
        start_time = time.time()
        if row['output'] != '':
            completion_embedding = OpenAIEmbeddings(
            ).embed_query(row['output'])
            highlight_embedding = OpenAIEmbeddings(
            ).embed_query(row['ground_truth'])
            cosine_similarity = np.dot(completion_embedding, highlight_embedding) / (
                np.linalg.norm(completion_embedding) * np.linalg.norm(highlight_embedding))
        else:
            cosine_similarity = 0
        end_time = time.time()
        latency = end_time - start_time
        scores.append((cosine_similarity, latency))
    # Add the scores to new columns in the DataFrame
    df_with_scores = df.copy()
    df_with_scores['cosine_similarity_openAI'] = [i[0] for i in scores]
    df_with_scores['latency_openAI'] = [i[1] for i in scores]
    return df_with_scores

# %%


def get_rouge_score(df):
    scores = []
    for i, row in df.iterrows():
        start_time = time.time()
        rouge_score = rouge.get_scores(row['output'], row['ground_truth'])
        end_time = time.time()
        latency = end_time - start_time
        scores.append((rouge_score, latency))
    # Add the scores to new columns in the DataFrame
    # rouge_score [i]['rouge-1']['f'] + rouge_score [i]['rouge-2']['f'] + rouge_score [i]['rouge-l']['f']
    df['rouge_score'] = [i[0] for i in scores]
    df['latency_rouge'] = [i[1] for i in scores]
    return df

# %%

def llm_eval_recall(df: pd.DataFrame) -> pd.DataFrame:
    """
    Evaluates whether outputs use all the data provided in the input values. Uses LLM for evaluation. Returns binary 0 / 1 assessment of each output
    """
    eval_prompt = PromptTemplate(
        template="""You are a professor grading a test. Your student was given an instruction and input values. The student then produced an answer. Here is the data:

[BEGIN DATA]
### INSTRUCTION:
{prompt_prefix}

### INPUT VALUES:
{input_values}

### STUDENT ANSWER:
{output}

###
[END DATA]

Did the student's answer use all the data provided in the input values? Work it out in a detailed step by step manner to be sure you have the right answer. Think step-by-step, then provide your explanations and final answer as a JSON instance that conforms to the JSON schema below. For example, the object {{"foo": "bar"}} conforms to the schema {{"foo": {{"description": "a string field", "type": "string"}}}}.
Here is the output schema:
```
{{"analysis": {{"description": "detailed review of the key data provided in the input values and if they were used in the student's answer"}}, "type": "string"}},
"final_answer": {{"description": "based on the analysis above, final answer as 'YES' or 'NO' if the student's answer used all the data provided in the input values", "type": "string", "enum": ["YES", "NO"]}}}}

Only respond with your answer as a JSON instance.

JSON OUTPUT:""",
        input_variables=['prompt_prefix', 'input_values', 'output'])

    eval_model = OpenAI(model_name="text-davinci-003", temperature=0.4, max_tokens=500)
    llm_scores = []
    for index, row in df.iterrows():
        score = 0
        if row['output'] != '':
            input_values = '\n'.join(['<' + input_var + '>: {' + input_val + '}' for input_var, input_val in row['input_values'].items()])
            formatted_eval_prompt = eval_prompt.format(prompt_prefix = row['prompt_prefix'],
                                                       input_values = input_values,
                                                       output = row['output'])
            if isinstance(eval_model, ChatOpenAI):
                formatted_eval_prompt = [HumanMessage(content=formatted_eval_prompt)]
            
            # Extract llm eval
            num_tries = 3
            for run in range(num_tries):
                print('LLM eval recall for prompt ID: ' + str(row['prompt_id']))
                response = eval_model.generate([formatted_eval_prompt]).generations[0][0].text.strip()
                try:
                    llm_eval = json.loads(response)
                    assert llm_eval['final_answer'] in ['YES', 'NO']
                    if llm_eval['final_answer'] == 'YES':
                        score = 1
                    else:
                        score = 0
                    break
                except:
                    continue
        llm_scores.append(score)

    # Add the scores to new columns in the DataFrame
    df_with_scores = df.copy()
    df_with_scores['llm_eval_recall'] = llm_scores
    return df_with_scores

# %%

def llm_eval_precision(df: pd.DataFrame) -> pd.DataFrame:
    """
    Evaluates whether outputs accurately use information from instruction and input values. Uses LLM for evaluation. Returns binary 0 / 1 assessment of each output
    """
    eval_prompt = PromptTemplate(
        template="""You are a professor grading a test. Your student was given an instruction and input values. The student then produced an answer. Here is the data:

[BEGIN DATA]
### INSTRUCTION:
{prompt_prefix}

### INPUT VALUES:
{input_values}

### STUDENT ANSWER:
{output}

###
[END DATA]

What new facts did the student's answer contain that was not previously provided in the instruction or input values? Check if any new facts, information, or assumptions are communicated. Work it out in a detailed step by step manner to be sure you have the right answer. Think step-by-step, then provide your explanations and final answer as a JSON instance that conforms to the JSON schema below. For example, the object {{"foo": "bar"}} conforms to the schema {{"foo": {{"description": "a string field", "type": "string"}}}}.
Here is the output schema:
```
{{"observation": {{"description": "detailed observation of any new facts in the student's answer that was not in the instruction or input values"}}, "type": "string"}},
"analysis": {{"description": "is the new information a clarification of the original instructions or does it contain new data that may not be true?", "type": "string"}},
"final_answer": {{"description": "based on the analysis above, final answer as 'YES' or 'NO' if the student's answer included new information beyond clarifications", "type": "string", "enum": ["YES", "NO"]}}}}

Only respond with your answer as a JSON instance.

JSON OUTPUT:""",
        input_variables=['prompt_prefix', 'input_values', 'output'])

    eval_model = OpenAI(model_name="text-davinci-003", temperature=0.4, max_tokens=500)
    llm_scores = []
    for index, row in df.iterrows():
        score = 0
        if row['output'] != '':
            input_values = '\n'.join(['<' + input_var + '>: {' + input_val + '}' for input_var, input_val in row['input_values'].items()])
            formatted_eval_prompt = eval_prompt.format(prompt_prefix = row['prompt_prefix'],
                                                       input_values = input_values,
                                                       output = row['output'])
            if isinstance(eval_model, ChatOpenAI):
                formatted_eval_prompt = [HumanMessage(content=formatted_eval_prompt)]
            
            # Extract llm eval
            num_tries = 3
            for run in range(num_tries):
                print('LLM eval precision for prompt ID: ' + str(row['prompt_id']))
                response = eval_model.generate([formatted_eval_prompt]).generations[0][0].text.strip()
                try:
                    llm_eval = json.loads(response)
                    assert llm_eval['final_answer'] in ['YES', 'NO']
                    if llm_eval['final_answer'] == 'YES':
                        score = 0
                    else:
                        score = 1
                    break
                except:
                    continue
        llm_scores.append(score)

    # Add the scores to new columns in the DataFrame
    df_with_scores = df.copy()
    df_with_scores['llm_eval_precision'] = llm_scores
    return df_with_scores

# %%

def llm_eval_ground_truth_comparison(df: pd.DataFrame) -> pd.DataFrame:
    """
    Evaluates whether outputs are similar to ground truth. Uses LLM for evaluation. Returns binary 0 / 1 assessment of each output
    """
    eval_prompt = PromptTemplate(
        template="""You are a professor grading a test. Your student was given an instruction and input values. The student then produced an output, which can be compared to an expert answer. Here is the data:

[BEGIN DATA]
==
EXPERT ANSWER:
{ground_truth}

==
STUDENT ANSWER:
{output}

==
[END DATA]

Is the student answer significantly different from the expert answer? Compare conceptual, stylistic, and formatting differences, including things like length of output. Work it out in a detailed step by step manner to be sure you have the right answer. Provide your step-by-step explanations and final answer as a JSON instance that conforms to the JSON schema below. For example, the object {{"foo": "bar"}} conforms to the schema {{"foo": {{"description": "a string field", "type": "string"}}}}.
Here is the output schema:
```
{{"observation": {{"description": "specific observations of conceptual, stylistic, and formatting differences between the student and expert answers"}}, "type": "string"}},
"analysis": {{"description": "are the differences between the student and expert answers significant?", "type": "string"}},
"final_answer": {{"description": "based on the analysis above, final answer as 'YES' or 'NO' if the student answer is significantly different from the expert answer", "type": "string", "enum": ["YES", "NO"]}}}}

Only respond with your answer as a JSON instance.

JSON OUTPUT:""",
        input_variables=['ground_truth', 'output'])

    eval_model = ChatOpenAI(model_name="gpt-3.5-turbo", max_tokens=500, openai_api_key=openai.api_key, temperature=0.4)
    llm_scores = []
    for index, row in df.iterrows():
        score = 0
        if row['output'] != '':
            formatted_eval_prompt = eval_prompt.format(ground_truth = row['ground_truth'],
                                                       output = row['output'])
            if isinstance(eval_model, ChatOpenAI):
                formatted_eval_prompt = [HumanMessage(content=formatted_eval_prompt)]
            
            # Extract llm eval
            num_tries = 3
            for run in range(num_tries):
                print('LLM eval ground truth comparison for prompt ID: ' + str(row['prompt_id']))
                response = eval_model.generate([formatted_eval_prompt]).generations[0][0].text.strip()
                try:
                    llm_eval = json.loads(response)
                    assert llm_eval['final_answer'] in ['YES', 'NO']
                    if llm_eval['final_answer'] == 'YES':
                        score = 0
                    else:
                        score = 1
                    break
                except:
                    continue
        llm_scores.append(score)

    # Add the scores to new columns in the DataFrame
    df_with_scores = df.copy()
    df_with_scores['llm_eval_ground_truth_comparison'] = llm_scores
    return df_with_scores

# %%

def llm_eval_style_check(df: pd.DataFrame) -> pd.DataFrame:
    """
    Evaluates whether output has appropriate content, style, and formatting. Uses LLM for evaluation. Returns binary 0 / 1 assessment of each output
    """
    eval_prompt = PromptTemplate(
        template="""You are a professor grading a test. Your student was given an instruction and input values. The student then produced an answer. Here is the data:

[BEGIN DATA]
### INSTRUCTION:
{prompt_prefix}

### INPUT VALUES:
{input_values}

### STUDENT ANSWER:
{output}

###
[END DATA]

Did the student answer have appropriate tone, style, and formatting? Work it out in a detailed step by step manner to be sure you have the right answer. Provide your step-by-step explanations and final answer as a JSON instance that conforms to the JSON schema below. For example, the object {{"foo": "bar"}} conforms to the schema {{"foo": {{"description": "a string field", "type": "string"}}}}.
Here is the output schema:
```
{{"analysis": {{"description": "detailed review of the tone, style, and formatting in the student answer"}}, "type": "string"}},
"feedback": {{"description": "based on the review, what are specific changes the student should make to their answer?"}}, "type": "string"}},
"final_answer": {{"description": "based on the analysis above, final answer as 'YES' or 'NO' if the student's answer has appropriate tone, style, and formatting", "type": "string", "enum": ["YES", "NO"]}}}}

Only respond with your answer as a JSON instance.

JSON OUTPUT:""",
        input_variables=['prompt_prefix', 'input_values', 'output'])

    eval_model = ChatOpenAI(model_name="gpt-3.5-turbo", max_tokens=500, openai_api_key=openai.api_key, temperature=0.4)
    llm_scores = []
    for index, row in df.iterrows():
        score = 0
        if row['output'] != '':
            input_values = '\n'.join(['<' + input_var + '>: {' + input_val + '}' for input_var, input_val in row['input_values'].items()])
            formatted_eval_prompt = eval_prompt.format(prompt_prefix = row['prompt_prefix'],
                                                       input_values = input_values,
                                                       output = row['output'])
            if isinstance(eval_model, ChatOpenAI):
                formatted_eval_prompt = [HumanMessage(content=formatted_eval_prompt)]
            
            # Extract llm eval
            num_tries = 3
            for run in range(num_tries):
                print('LLM eval style for prompt ID: ' + str(row['prompt_id']))
                response = eval_model.generate([formatted_eval_prompt]).generations[0][0].text.strip()
                try:
                    llm_eval = json.loads(response)
                    assert llm_eval['final_answer'] in ['YES', 'NO']
                    if llm_eval['final_answer'] == 'YES':
                        score = 1
                    else:
                        score = 0
                    break
                except:
                    continue
        llm_scores.append(score)

    # Add the scores to new columns in the DataFrame
    df_with_scores = df.copy()
    df_with_scores['llm_eval_style'] = llm_scores
    return df_with_scores

# %%


def run_evaluation(df):
    """
    Add evaluation score metrics to dataframe
    """
    # Add cosine similarity scores
    df_with_scores = get_semantic_cosine_similarity_openAI(df)
    # df_with_scores['avg_score'] = df_with_scores[[
    #     'cosine_similarity_openAI']].mean(axis=1)
    
    return df_with_scores

# %%

def llm_eval_shortlist(evaluation_df: pd.DataFrame, num_of_top_rows_overall: int, file_name = str) -> pd.DataFrame:
    """
    Shortlists a given set of prompt candidates using LLM evaluation on the inference results. Returns a DataFrame with the shortlisted prompts only.
    """
    # Apply LLM evals
    # scores_df = llm_eval_recall(evaluation_df)
    scores_df = llm_eval_precision(evaluation_df)
    scores_df = llm_eval_ground_truth_comparison(scores_df)
    # scores_df = llm_eval_style_check(evaluation_df)

    # Calculate column with total score across LLM eval metrics
    scores_df['llm_eval_total_score'] = (scores_df['llm_eval_precision'] + scores_df['llm_eval_ground_truth_comparison'])
    
    # Log results
    update_report_llm_eval(scores_df, file_name)

    # Group the DataFrame by prompt_id and generation_id, then compute average total llm eval score
    df = scores_df.groupby(['prompt_id', 'generation_id']).agg({
        'llm_eval_total_score': ['mean'],
        'cosine_similarity_openAI': ['mean'],
        'latency': ['mean'],
        'prompt_object': 'first',
        'prompt_prefix': 'first',
        'model_object': 'first',
    }).reset_index()
    df.columns = ['_'.join(col).strip('_') for col in df.columns.values]
    df = df.rename(columns={'prompt_object_first': 'prompt_object',
                            'prompt_prefix_first': 'prompt_prefix',
                            'model_object_first': 'model_object'})

    # Sort the DataFrame by the llm_eval_total_score_mean and cosine_similarity_openAI_mean in descending order
    sorted_df = df.sort_values(['llm_eval_total_score_mean', 'cosine_similarity_openAI_mean'],
                               ascending=[False, False])

    # Get the top n rows
    result_df = sorted_df.head(num_of_top_rows_overall)

    # Return shortlisted prompt candidates. Keep subset of columns, update generation ID, and reset index to start at 0
    result_df = result_df[['prompt_id', 'generation_id', 'prompt_object', 'prompt_prefix', 'model_object']]
    result_df = result_df.reset_index(drop=True)
    return result_df

# %%


def prompt_shortlist(df, num_of_top_rows_overall, num_of_top_rows_per_metric, metric):
    # Check if the metric is in the DataFrame
    if num_of_top_rows_per_metric > 0 and metric not in df.columns:
        return "Error: Metric not found in the DataFrame"

    # Group the DataFrame by prompt_id, generation_id
    df = df.groupby(['prompt_id', 'generation_id']).agg({
        'cosine_similarity_openAI': ['min', 'max', 'mean', ('median', lambda x: x.median())],
        'prompt_object': 'first',
        'prompt_prefix': 'first',
        'model_object': 'first',
    }).reset_index()

    df.columns = ['_'.join(col).strip('_') for col in df.columns.values]
    df = df.rename(columns={'prompt_object_first': 'prompt_object',
                            'prompt_prefix_first': 'prompt_prefix',
                            'model_object_first': 'model_object'})
    
    # Add score that computes average of min and mean cosine similarity value
    df['min_mean_cosine_similarity'] = (df['cosine_similarity_openAI_min'] + df['cosine_similarity_openAI_mean']) / 2

    # Sort the DataFrame by the cosine similarity score in descending order
    sorted_df = df.sort_values('min_mean_cosine_similarity', ascending=False)

    # Get the top n rows with highest score overall
    top_n_overall = sorted_df.head(num_of_top_rows_overall)

    # drop top_n_overall from sorted_df
    sorted_df = sorted_df.drop(top_n_overall.index)

    # Get the top m rows with highest score for each unique metric
    top_m_per_metric = sorted_df.groupby(metric).apply(
        lambda x: x.nlargest(num_of_top_rows_per_metric, 'min_mean_cosine_similarity'))
    top_m_per_metric = top_m_per_metric.reset_index(
        drop=True)  # reset index to start at 0

    # Combine the two DataFrames and sort by score in descending order
    result_df = pd.concat([top_n_overall, top_m_per_metric]).sort_values(
        'min_mean_cosine_similarity', ascending=False)
    result_df = result_df.reset_index(drop=True)  # reset index to start at 0
    return result_df

# %% [markdown]
# **Report Generation**

# %%

def apply_prompt_object(x):
    if isinstance(x[0], FewShotPromptTemplate):
        template_string = x[0].prefix
        input_variables = json.dumps(x[0].input_variables)
    else:
        template_string = x[0].template
        input_variables = json.dumps(x[0].input_variables)
    return pd.Series([template_string, input_variables])

# %%


def apply_model_object(x):
    if isinstance(x[0], ChatOpenAI):
        model_name = x[0].model_name
        temperature = x[0].model_kwargs['temperature']
    elif isinstance(x[0], OpenAI):
        model_name = x[0].model_name
        temperature = x[0].temperature
    elif isinstance(x[0], HuggingFaceHub):
        model_name = x[0].repo_id
        temperature = x[0].model_kwargs['temperature']
    return pd.Series([model_name, temperature])

# %%
# Function to extract attributes into separate columns


def extract_prompt_attributes(prompt):
    return pd.Series({'input_variables': prompt.input_variables,
                      'template': prompt.template})

# %%


def update_report(df, file_name, stage_id):
    if len(df) == 0:
        return
    
    # Make a copy of the dataframe to avoid modifying the original dataframe
    df_copy = df.copy()
    # print(df)

    # add stage_id to the dataframe copy
    df_copy['stage'] = stage_id

    # extract input variables into separate columns
    df_copy[['input_variables']
            ] = df_copy['prompt_object'].apply(lambda x: pd.Series([json.dumps(x[0].input_variables)]))
    df_copy[['model_name', 'temperature']
            ] = df_copy['model_object'].apply(apply_model_object)

    # drop original prompt_object column
    df_copy.drop('prompt_object', axis=1, inplace=True)
    df_copy.drop('model_object', axis=1, inplace=True)

    # Convert dictionary values to strings for specific columns
    cols_to_convert = ['prompt_id', 'generation_id', 'input_values', 'output',
                       'ground_truth', 'stage', 'prompt_prefix', 'input_variables', 'model_name']
    df_copy[cols_to_convert] = df_copy[cols_to_convert].applymap(
        lambda x: json.dumps(x) if isinstance(x, dict) else x)

    # Create or open the excel file
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
    else:
        wb = Workbook()
        # add a new sheet named 'results' to the workbook
        wb.create_sheet('results')
        # add the df headers as column headers to the 'results' sheet
        wb['results'].append(list(df_copy.columns))

    # write the dataframe copy to the 'results' sheet
    ws = wb['results']
    for r in dataframe_to_rows(df_copy, index=False, header=False):
        ws.append(r)

    # Save and close the workbook
    wb.save(file_name)
    wb.close()

# %%

def update_report_llm_eval(df: pd.DataFrame, file_name: str):
    """
    Log results from LLM eval
    """
    if len(df) == 0:
        return
    
    # Make a copy of the dataframe to avoid modifying the original dataframe
    df_copy = df.copy()
    # print(df)

    # extract input variables into separate columns
    df_copy[['input_variables']
            ] = df_copy['prompt_object'].apply(lambda x: pd.Series([json.dumps(x[0].input_variables)]))
    df_copy[['model_name', 'temperature']
            ] = df_copy['model_object'].apply(apply_model_object)

    # drop original prompt_object column
    df_copy.drop('prompt_object', axis=1, inplace=True)
    df_copy.drop('model_object', axis=1, inplace=True)

    # Convert dictionary values to strings for specific columns
    cols_to_convert = ['prompt_id', 'generation_id', 'input_values', 'output',
                       'ground_truth', 'prompt_prefix', 'input_variables', 'model_name']
    df_copy[cols_to_convert] = df_copy[cols_to_convert].applymap(
        lambda x: json.dumps(x) if isinstance(x, dict) else x)

    # Create or open the excel file
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
    else:
        wb = Workbook()
    # add a new sheet named 'llm_eval_results' to the workbook
    wb.create_sheet('llm_eval_results')
    # add the df headers as column headers to the 'results' sheet
    wb['llm_eval_results'].append(list(df_copy.columns))

    # write the dataframe copy to the 'results' sheet
    ws = wb['llm_eval_results']
    for r in dataframe_to_rows(df_copy, index=False, header=False):
        ws.append(r)

    # Save and close the workbook
    wb.save(file_name)
    wb.close()

# %%


def finalize_report(file_name, prompt_candidates: pd.DataFrame):

    df = pd.read_excel(file_name, sheet_name='results')

    df_prompt = df.groupby(['prompt_id', 'generation_id', 'stage']).agg(
        {'cosine_similarity_openAI': ['min', 'max', 'mean', ('median', lambda x: x.median())]})
    df_prompt[('cosine_similarity_openAI', 'min_mean')] = (df_prompt[('cosine_similarity_openAI', 'min')] + df_prompt[('cosine_similarity_openAI', 'mean')]) / 2
    df_prompt = df_prompt.unstack(level=-1)
    df_prompt = df_prompt.stack(level=0).swaplevel(
        0, 1, axis=1).sort_index(axis=1)

    df_generation = df.groupby(['generation_id', 'stage']).agg(
        {'cosine_similarity_openAI': ['min', 'max', 'mean', ('median', lambda x: x.median())]})
    df_generation = df_generation.unstack(level=-1)
    df_generation = df_generation.stack(level=0).swaplevel(
        0, 1, axis=1).sort_index(axis=1)

    # load existing workbook
    book = load_workbook(file_name)

    # create new sheet
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book

    df_prompt.to_excel(writer, sheet_name='prompt_summary')

    df_generation.to_excel(writer, sheet_name='generation_summary')

    # extract template string and input variables into separate columns
    prompt_candidates[['input_variables']
                      ] = prompt_candidates['prompt_object'].apply(lambda x: pd.Series([json.dumps(x[0].input_variables)]))
    prompt_candidates[['model_name', 'temperature']
                      ] = prompt_candidates['model_object'].apply(apply_model_object)
    prompt_candidates.to_excel(writer, sheet_name='final_results')

    # save the changes to the workbook
    writer.save()

# %% [markdown]
# MAIN


# %%
# Load evaluation datasets
csvfile = open('./Evaluation_Datasets/usecase_Sentiment.csv', newline='')
evaluation_dataset_sentiment = pd.DataFrame(list(csv.reader(csvfile))[1:])

csvfile = open('./Evaluation_Datasets/usecase_AdversarialQA.csv', newline='')
evaluation_dataset_QA = pd.DataFrame(list(csv.reader(csvfile))[1:])

# csvfile = open(
#     './Evaluation_Datasets/usecase_email_cta_updated.csv', newline='')
# evaluation_dataset_email_cta = pd.DataFrame(list(csv.reader(csvfile))[1:])

csvfile = open('./Evaluation_Datasets/usecase_Summarization.csv', newline='')
evaluation_dataset_summarization = pd.DataFrame(list(csv.reader(csvfile))[1:])


# %%

def run_algorithm(experiment: dict, selected_model: tuple, file_name: str, num_prompts: int, global_prompt_id: list) -> pd.DataFrame:
    """
    Returns optimal prompt object and configured model object for selected model
    """

    # Stage 1 - Prompt Generation
    prompt_candidates_user_objective = prompt_generation_user_objective(
        experiment, selected_model, num_prompts, global_prompt_id)
    print("completed prompt generation user objective")

    prompt_candidates_role_play = prompt_generation_pattern_role_play(
        experiment, selected_model, num_prompts, global_prompt_id)
    print("completed prompt generation role play")

    prompt_candidates_user_objective_training_data = prompt_generation_user_objective_training_data(
        experiment, selected_model, num_prompts, global_prompt_id)
    print("completed prompt generation user objective with training data")

    prompt_candidates_first_pass = pd.concat(
        [prompt_candidates_user_objective, prompt_candidates_role_play, prompt_candidates_user_objective_training_data],
        ignore_index=True).reset_index(drop=True)
    prompt_candidates_variants = prompt_generation_variants(
        experiment, prompt_candidates_first_pass, 3, global_prompt_id)
    prompt_candidates_first_pass = pd.concat(
        [prompt_candidates_first_pass, prompt_candidates_variants],
        ignore_index=True).reset_index(drop=True)
    print("completed prompt generation variants")

    prompt_candidates_first_pass = cluster_shortlist(
        prompt_candidates_first_pass, min(10, len(prompt_candidates_first_pass)))
    print("completed cluster shortlisting")

    inference_first_pass = run_inference(
        experiment, prompt_candidates_first_pass, "test")
    print("completed inference")
    evaluation_first_pass = run_evaluation(inference_first_pass)
    print("completed evaluation")
    update_report(evaluation_first_pass, file_name,
                  'prompt_generation_stage_01')
    print("completed update report")
    shortlist_first_pass = prompt_shortlist(
        evaluation_first_pass, 5, 0, 'generation_id')
    print("completed shortlist")

    # Stage 2 - Few Shots
    if isinstance(selected_model[0], OpenAI) and (selected_model[0].model_name == "text-babbage-001"):
        num_few_shots = 2
    elif isinstance(selected_model[0], HuggingFaceHub) and (selected_model[0].repo_id == 'google/flan-ul2'):
        num_few_shots = 2
    else:
        num_few_shots = 5
    prompt_candidates_second_pass = few_shots(
        experiment, shortlist_first_pass, num_few_shots, global_prompt_id)
    print("completed few shots")
    inference_second_pass = run_inference(
        experiment, prompt_candidates_second_pass, "test")
    print("completed inference few shots")
    evaluation_second_pass = run_evaluation(inference_second_pass)
    print("completed evaluation few shots")
    update_report(evaluation_second_pass, file_name,
                  'prompt_generation_stage_few_shots_02')
    print("completed update report few shots")
    concatenated_evaluation_second_pass = pd.concat(
        [evaluation_second_pass, evaluation_first_pass], ignore_index=True).reset_index(drop=True)
    shortlist_second_pass = prompt_shortlist(
        concatenated_evaluation_second_pass, 3, 0, 'generation_id')
    filtered_evaluation_second_pass = concatenated_evaluation_second_pass[
        concatenated_evaluation_second_pass.prompt_id.isin(shortlist_second_pass.prompt_id)
    ]
    print("completed shortlist")

    # Stage 3 - Prompt Improvement TIM
    prompt_candidates_third_pass = prompt_generation_targeted_improvement(
        experiment, shortlist_second_pass, global_prompt_id)
    print("completed targeted improvement TIMIII")
    inference_third_pass = run_inference(
        experiment, prompt_candidates_third_pass, "test")
    print("completed inference TIM")
    evaluation_third_pass = run_evaluation(inference_third_pass)
    print("completed evaluation TIM")
    update_report(evaluation_third_pass, file_name, 'prompt_TIM_stage_03')
    print("completed update report")
    concatenated_evaluation_third_pass = pd.concat(
        [filtered_evaluation_second_pass, evaluation_third_pass], ignore_index=True).reset_index(drop=True)
    shortlist_third_pass = prompt_shortlist(
        concatenated_evaluation_third_pass, 3, 0, 'generation_id')
    print("completed shortlist TIM")
    filtered_evaluation_third_pass = concatenated_evaluation_third_pass[
        concatenated_evaluation_third_pass.prompt_id.isin(shortlist_third_pass.prompt_id)
    ]

    # Stage 4 - LLM eval shortlist
    shortlist_fourth_pass = llm_eval_shortlist(filtered_evaluation_third_pass, 1, file_name)
    print("completed LLM eval shortlist")

    # Stage 5 - Temperature variation
    prompt_candidates_fifth_pass = prompt_generation_temperature_variation(
        experiment, shortlist_fourth_pass, 4, global_prompt_id)
    inference_fifth_pass = run_inference(
        experiment, prompt_candidates_fifth_pass, "test")
    evaluation_fifth_pass = run_evaluation(inference_fifth_pass)
    update_report(evaluation_fifth_pass, file_name,
                  'prompt_generation_stage_temp_variation_05')
    shortlist_fifth_pass = prompt_shortlist(
        evaluation_fifth_pass, 1, 0, 'generation_id')

    return shortlist_fifth_pass

# %%


def initialize_horizon() -> dict:
    """
    Gathers input from user and returns dictionary with relevant details for experiment
    """
    print(
        '\nSelect from one of the following tasks: [sentiment analysis, text summarization, text generation, reasoning, other, skip]\n')
    task_type = input('Enter your task type: ')

    if task_type == 'email-cta':
        print('Defaulting to email-cta task')
        task_type = 'Text generation'
        user_objective = 'predict the call to action for the given email'
        evaluation_dataset_path = './Evaluation_Datasets/usecase_email_cta_updated.csv'

    elif task_type == 'summary':
        print('Defaulting to summary task')
        task_type = 'Text summarization'
        user_objective = 'summarize the given text'
        evaluation_dataset_path = './Evaluation_Datasets/usecase_Summarization.csv'

    elif task_type == "sell_scale":
        print('Defaulting to sell_scale task')
        task_type = 'Text generation'
        user_objective = 'Generate the first 1-3 lines of a personalized email to a prospect from the perspective of a sales person at a security tech company'
        evaluation_dataset_path = './Evaluation_Datasets/usecase_sell_scale_v1.csv'
        
    else:
        user_objective = input(
            'Please provide a detailed description of your objective: ')
        evaluation_dataset_path = input(
            'Please enter the path to your evaluation dataset with at least 15 rows: ')
    
    print('Thank you! Preparing your optimal prompt, model, and model parameters. Check back in a few mins :)')
    csvfile = open(evaluation_dataset_path, newline='')
    file_lines = list(csv.reader(csvfile))
    input_variables = file_lines[0][:-1]
    evaluation_dataset = pd.DataFrame(file_lines[1:])

    # Shuffle rows
    evaluation_dataset = evaluation_dataset.sample(frac=1).reset_index(drop=True)

    input_values_test = [evaluation_dataset.iloc[0:10, i].reset_index(
        drop=True) for i in range(len(input_variables))]
    ground_truth_test = evaluation_dataset.iloc[0:10, -1].reset_index(drop=True)
    input_values_train = [evaluation_dataset.iloc[10:15, i].reset_index(
        drop=True) for i in range(len(input_variables))]
    ground_truth_train = evaluation_dataset.iloc[10:15, -1].reset_index(drop=True)
    input_values_few_shots = [evaluation_dataset.iloc[10:15, i].reset_index(
        drop=True) for i in range(len(input_variables))]
    ground_truth_few_shots = evaluation_dataset.iloc[10:15, -1].reset_index(drop=True)

    experiment = {
        'task_type': task_type,
        'user_objective': user_objective,
        'input_variables': input_variables,
        'evaluation_dataset': evaluation_dataset,
        'input_values_test': input_values_test,
        'ground_truth_test': ground_truth_test,
        'input_values_train': input_values_train,
        'ground_truth_train': ground_truth_train,
        'input_values_few_shots': input_values_few_shots,
        'ground_truth_few_shots': ground_truth_few_shots,
    }
    return experiment

# %%


class Task:
    def __init__(self, prompt_object: tuple, model_object: tuple, prompt_id: int, file_name: str) -> None:
        """
        Creates a new Task object that uses the provided prompt_object, model_object, and associated input variables.
        prompt_object and model_object must be passed in as a tuple with a single element
        """
        self.prompt_object = prompt_object[0]
        self.model_object = model_object[0]
        self.input_variables = self.prompt_object.input_variables
        self.prompt_id = prompt_id
        self.file_name = file_name
        self.results_df = pd.read_excel(file_name, sheet_name='results')
        self.final_results_df = pd.read_excel(file_name, sheet_name='final_results')

    def generate(self, **input_values) -> str:
        """
        Generate output using the provided prompt_object and model_object.
        Arguments must be passed in as a key-value pair for each input variable
        """
        prompt = self.prompt_object.format(**input_values)

        if type(self.model_object) == ChatOpenAI:
            prompt = [HumanMessage(content=prompt)]
        output = self.model_object.generate(
            [prompt]).generations[0][0].text.strip()

        return output
    
    def display_details(self, verbose=False):
        """
        Prints details about Task object and evaluation job.
        """
        total_evaluations_considered = len(self.results_df)
        accuracy_score = float(self.final_results_df['min_mean_cosine_similarity'].loc[self.final_results_df['prompt_id'] == self.prompt_id][0])
        average_latency = self.results_df['latency'].loc[self.results_df['prompt_id'] == self.prompt_id].mean()
        few_shot = type(self.prompt_object) is FewShotPromptTemplate
        prompt_prefix = self.final_results_df['prompt_prefix'].loc[self.final_results_df['prompt_id'] == self.prompt_id][0]
        cleaned_prompt_prefix = []
        for line in prompt_prefix.split('\n'):
            if len(line) == 0:
                cleaned_prompt_prefix.append(line)
            else:
                cleaned_prompt_prefix += [line[i:i+100].strip() for i in range(0, len(line), 100)]
        cleaned_prompt_prefix = '\n'.join(cleaned_prompt_prefix)

        test_input_values = dict((k, v) for k, v in zip(self.input_variables, ['test'] * len(self.input_variables)))
        example_prompt_string = self.prompt_object.format(**test_input_values)
        cleaned_example_prompt_string = []
        for line in example_prompt_string.split('\n'):
            if len(line) == 0:
                cleaned_example_prompt_string.append(line)
            else:
                cleaned_example_prompt_string += [line[i:i+100].strip() for i in range(0, len(line), 100)]
        cleaned_example_prompt_string = '\n'.join(cleaned_example_prompt_string)

        model_name = self.final_results_df['model_name'].loc[self.final_results_df['prompt_id'] == self.prompt_id][0]
        model_temperature = float(self.final_results_df['temperature'].loc[self.final_results_df['prompt_id'] == self.prompt_id][0])

        if verbose == False:
            print(tabulate([
                ['Total evaluations considered', total_evaluations_considered],
                ['Performance score', f'{accuracy_score: .2f}'],
                ['Average latency', f'{average_latency: .2f}'],
                ['Is few shot?', few_shot],
                ['Prompt prefix', cleaned_prompt_prefix]
            ],
            headers=['Parameter', 'Value'],
            tablefmt='grid'))
        elif verbose == True:
            print(tabulate([
                ['Total evaluations considered', total_evaluations_considered],
                ['Performance score', f'{accuracy_score: .2f}'],
                ['Average latency (s)', f'{average_latency: .2f}'],
                ['Is few shot?', few_shot],
                ['Prompt prefix', cleaned_prompt_prefix],
                ['Example full prompt string', cleaned_example_prompt_string],
                ['Model', model_name],
                ['Model temperature', f'{model_temperature: .2f}']
            ],
            headers=['Parameter', 'Value'],
            tablefmt='grid'))
        return
        

# %%
# Main


def create_task() -> Task:
    file_name = 'Horizon_email_gen_v2.xlsx'
    num_prompts = 5
    global_prompt_id = [0]

    experiment = initialize_horizon()
    selected_models = [
        # (HuggingFaceHub(repo_id="google/flan-ul2", model_kwargs={"temperature":0.7, "max_length":10000},
        #                 huggingfacehub_api_token=HUGGINGFACEHUB_API_TOKEN), ),
        (ChatOpenAI(model_name="gpt-3.5-turbo", max_tokens=500,
         openai_api_key=openai.api_key, temperature=0.7),),
        # (OpenAI(model_name="text-davinci-003", max_tokens=500),),
        # (OpenAI(model_name="text-babbage-001", max_tokens=200),)
    ]

    # Generate top candidate for each model
    selected_candidates = pd.DataFrame()
    for model in selected_models:
        if type(model[0]) in [OpenAI, ChatOpenAI]:
            print('Working on model ' + model[0].model_name)
        elif type(model[0]) in [HuggingFaceHub]:
            print('Working on model ' + model[0].repo_id)
        candidate = run_algorithm(
            experiment, model, file_name, num_prompts, global_prompt_id)
        selected_candidates = pd.concat(
            [selected_candidates, candidate], ignore_index=True).reset_index(drop=True)

    # Sort selected candidates and select final candidate
    sorted_candidates = selected_candidates.sort_values(
        'min_mean_cosine_similarity', ascending=False)
    finalize_report(file_name, sorted_candidates)
    final_candidate = sorted_candidates.head(1)

    task = Task(prompt_object=final_candidate['prompt_object'].iloc[0],
                model_object=final_candidate['model_object'].iloc[0],
                prompt_id=final_candidate['prompt_id'].iloc[0],
                file_name=file_name)
    return task

# %%

def create_task_manually(task_type: str) -> Task:

    def get_example_dataset(experiment: dict) -> list:
        examples = []
        for i in range(len(experiment['ground_truth_train'])):
            example = {}
            for j in range(len(experiment['input_variables'])):
                example[experiment['input_variables'][j]
                        ] = experiment['input_values_train'][j].iloc[i]
            example['output'] = experiment['ground_truth_train'].iloc[i]
            examples.append(example)
        return examples

    def create_few_shot_prompt(experiment: dict, prompt_prefix: str) -> FewShotPromptTemplate:
        examples = get_example_dataset(experiment)
        few_shot_prompt_prefix = prompt_prefix + "\n\n==\nEXAMPLES:"
        # iterate through each input variable
        exampleFormatterTemplate = ""
        for j in range(len(experiment['input_variables'])):
            exampleFormatterTemplate += "<" + \
                experiment['input_variables'][j] + \
                ">: {" + experiment['input_variables'][j] + "}\n"

        suffix_request = "==\nBEGIN:\n\n" + exampleFormatterTemplate + "<OUTPUT>:"
        # add the output variable
        exampleFormatterTemplate += "<OUTPUT>: {output}"
        new_input_variables = experiment['input_variables'] + ['output']
        # create the prompt template
        example_prompt = PromptTemplate(
            input_variables=new_input_variables,
            template=exampleFormatterTemplate,
        )

        # create the example selector
        # examples: This is the list of examples available to select from.
        # embeddings: This is the Embeddings class that is used to embed the examples.
        # vector_store: This is the VectorStore class that is used to store the embeddings and do a similarity search over.
        # k: This is the number of examples to produce.
        example_selector = MaxMarginalRelevanceExampleSelector.from_examples(
            examples, OpenAIEmbeddings(), FAISS, k=5)

        # create the few shot prompt template
        few_shot_prompt = FewShotPromptTemplate(
            example_selector=example_selector,
            example_prompt=example_prompt,
            prefix=few_shot_prompt_prefix,
            suffix=suffix_request,
            input_variables=experiment['input_variables'],
        )
        return few_shot_prompt

    if task_type == 'email_gen':
        print('Preparing endpoint for email generation task')
        csvfile = open(
            './Evaluation_Datasets/usecase_sell_scale_v1_copy.csv', newline='')
        evaluation_dataset_sell_scale = pd.DataFrame(
            list(csv.reader(csvfile))[1:])
        experiment = {
            'task_type': 'text generation',
            'user_objective': 'Generate the first 1-3 lines of a personalized email to a prospect from the perspective of a sales person at a security tech company',
            'input_variables': ['name', 'industry', 'company', 'title', 'notes'],
            'evaluation_dataset': evaluation_dataset_sell_scale,
            'input_values_test': [
                evaluation_dataset_sell_scale.iloc[0:10, 0].reset_index(drop=True),
                evaluation_dataset_sell_scale.iloc[0:10, 1].reset_index(drop=True),
                evaluation_dataset_sell_scale.iloc[0:10, 2].reset_index(drop=True),
                evaluation_dataset_sell_scale.iloc[0:10, 3].reset_index(drop=True),
                evaluation_dataset_sell_scale.iloc[0:10, 4].reset_index(drop=True),
            ],
            'ground_truth_test': evaluation_dataset_sell_scale.iloc[0:10, 5].reset_index(drop=True),
            'input_values_train': [
                evaluation_dataset_sell_scale.iloc[0:15, 0].reset_index(drop=True),
                evaluation_dataset_sell_scale.iloc[0:15, 1].reset_index(drop=True),
                evaluation_dataset_sell_scale.iloc[0:15, 2].reset_index(drop=True),
                evaluation_dataset_sell_scale.iloc[0:15, 3].reset_index(drop=True),
                evaluation_dataset_sell_scale.iloc[0:15, 4].reset_index(drop=True),
            ],
            'ground_truth_train': evaluation_dataset_sell_scale.iloc[0:15, 5].reset_index(drop=True)
        }
        prompt_prefix = "Write the first 1-3 lines of a personalized email to a prospect from the perspective of a salesperson at a company in the <industry> industry. Make sure to include the company name, the prospect's title, and any notes that you may have about the prospect or the company. Use the input variables to craft a unique email that highlights the strengths of the company and the prospect's role."
        prompt = create_few_shot_prompt(experiment, prompt_prefix)
        model = OpenAI(model_name="text-davinci-003", temperature=0.37, max_tokens=500)
        file_name = './email_gen_demo.xlsx'
        task = Task(prompt_object=(prompt,), model_object=(model,), prompt_id=77, file_name=file_name)
        return task
    else:
        raise ValueError('No standard task for the given input')
    
# %%

def get_evaluation_data(file_name: str) -> dict:
    """
    Checks that evaluation data meets minimum data requirements and is properly formatted.
    If all checks pass, then returns a dict with the evaluation dataset as DataFrame and other info

    Inputs:
    file_name: str -> path to evaluation dataset

    Outputs:
    evaluation_data_configuration: dict -> contains information about evaluation data contents (e.g., tokens, characters) and applicable LLMs
    """
    evaluation_data_configuration = {}

    # Check that evaluation data is at most 1 MB file size
    if os.path.getsize(file_name) > 1000000:
        raise AssertionError('Evaluation dataset can be at most 1 MB large.')
    
    # Try to import evaluation dataset
    try:
        csvfile = open(file_name, newline='')
        data = list(csv.reader(csvfile))
    except:
        raise AssertionError('Could not import evaluation data. Make sure to upload in csv format.')

    # Check that there are at least 15 rows of data
    if len(data) < 16:
        raise AssertionError('There must be at least 15 rows of evaluation data.')
    
    # Check that there is at least 1 column. Last column is assumed to be the ground truth
    columns = data[0]
    if len(columns) < 1:
        raise AssertionError('There must be at least 1 column. The rightmost column is assumed to be the ground truth.')
    
    # Check that each input variable is a single string with letters, numbers, and underscores only (no spaces)
    input_variables = columns[:-1]
    for input_var in input_variables:
        if not re.match(r'^[A-Za-z0-9_]+$', input_var):
            raise AssertionError('Input variable names must be a single string letters, numbers, and underscores only (no spaces allowed)')
    
    # Get max data length used for input values and ground truth across all rows
    evaluation_dataset = pd.DataFrame(data[1:]).sample(frac=1).reset_index(drop=True) # Shuffle dataset
    encoding_turbo = tiktoken.encoding_for_model("gpt-3.5-turbo")
    encoding_davinci = tiktoken.encoding_for_model("text-davinci-003")
    max_input_tokens = 0
    max_ground_truth_tokens = 0
    max_input_characters = 0
    max_ground_truth_characters = 0
    for index, row in evaluation_dataset.iterrows():
        row_input_tokens = 0
        row_ground_truth_tokens = 0
        row_input_characters = 0
        row_ground_truth_characters = 0

        for input_value_index in range(len(input_variables)):
            row_input_tokens += max(len(encoding_turbo.encode(row.iloc[input_value_index])),
                                    len(encoding_davinci.encode(row.iloc[input_value_index])))
            row_input_characters += len(row[input_value_index])
        row_ground_truth_tokens = max(len(encoding_turbo.encode(row.iloc[-1])),
                                       len(encoding_davinci.encode(row.iloc[-1])))
        row_ground_truth_characters = len(row.iloc[-1])

        max_input_tokens = max(max_input_tokens, row_input_tokens)
        max_ground_truth_tokens = max(max_ground_truth_tokens, row_ground_truth_tokens)
        max_input_characters = max(max_input_characters, row_input_characters)
        max_ground_truth_characters = max(max_ground_truth_characters, row_ground_truth_characters)

    # Add data length from input and output variable names
    max_input_tokens += max(len(encoding_turbo.encode('\n'.join(input_variables) + '\n<OUTPUT>:')),
                            len(encoding_davinci.encode('\n'.join(input_variables) + '\n<OUTPUT>:')))
    max_input_characters += len('\n'.join(input_variables) + '\n<OUTPUT>:')

    evaluation_data_configuration = {
        'input_variables': input_variables,
        'evaluation_dataset': evaluation_dataset,
        'max_input_tokens': max_input_tokens,
        'max_ground_truth_tokens': max_ground_truth_tokens,
        'max_input_characters': max_input_characters,
        'max_ground_truth_characters': max_ground_truth_characters
    }

    # Get applicable LLMs based on context length
    applicable_llms = get_applicable_llm_config(evaluation_data_configuration)
    if len(applicable_llms) == 0:
        raise AssertionError('Input and output data length exceed context length of available LLMs (assumes data length for instruction string plus buffer).')
    
    # Check that text-davinci-003 is an applicable LLM and that at least 5 few shot examples fit (needed for prompt generation)
    if 'text-davinci-003' not in applicable_llms or applicable_llms['text-davinci-003']['max_few_shots'] < 3:
        raise AssertionError('Input and output data length exceed context length of available LLMs (assumes few shot examples are used).')

    evaluation_data_configuration['applicable_llms'] = applicable_llms

    # Segment evaluation data into test and training data (assumed 80/20 split, with at least 5 training data points)
    evaluation_data_configuration['num_train_data'] = max(5, int(len(evaluation_dataset) * 0.2))
    evaluation_data_configuration['num_test_data'] = len(evaluation_dataset) - evaluation_data_configuration['num_train_data']
    evaluation_data_configuration['input_data_train'] = evaluation_dataset.iloc[
        :evaluation_data_configuration['num_train_data'], :-1].reset_index(drop=True)
    evaluation_data_configuration['ground_truth_data_train'] = evaluation_dataset.iloc[
        :evaluation_data_configuration['num_train_data'], -1:].reset_index(drop=True)
    evaluation_data_configuration['input_data_test'] = evaluation_dataset.iloc[
        evaluation_data_configuration['num_train_data']:, :-1].reset_index(drop=True)
    evaluation_data_configuration['ground_truth_data_test'] = evaluation_dataset.iloc[
        evaluation_data_configuration['num_train_data']:, -1:].reset_index(drop=True)

    return evaluation_data_configuration

# %%

def get_applicable_llm_config(evaluation_data_configuration: dict) -> dict:
    """
    Takes in a dict that contains information about the evauation data (as returned by get_evaluation_data) and determines which LLMs are applicable,
    the associated number of few shot examples that can be used based on token and character limits, and max output token field

    Inputs:
    evaluation_data_configuration: dict -> contains information about evaluation data contents (e.g., tokens, characters)

    Outputs:
    applicable_llms: dict -> returns dict containing each applicable llm and associated config
    """
    # Determine data length of zero-shot prompt
    zero_shot_tokens = llm_data_assumptions['buffer_tokens'] + llm_data_assumptions['buffer_tokens'] + \
        (evaluation_data_configuration['max_input_tokens'] * llm_data_assumptions['input_output_multiplier'])
    zero_shot_characters = llm_data_assumptions['buffer_characters'] + llm_data_assumptions['buffer_characters'] + \
        (evaluation_data_configuration['max_input_characters'] * llm_data_assumptions['input_output_multiplier'])
    
    # Determine data length of few shot example
    max_few_shot_tokens = (evaluation_data_configuration['max_input_tokens'] + evaluation_data_configuration['max_ground_truth_tokens']) * \
        llm_data_assumptions['input_output_multiplier']
    max_few_shot_characters = evaluation_data_configuration['max_input_characters'] + evaluation_data_configuration['max_ground_truth_characters'] * \
        llm_data_assumptions['input_output_multiplier']
    
    # Determine max output data length
    max_output_tokens = int(evaluation_data_configuration['max_ground_truth_tokens'] * llm_data_assumptions['input_output_multiplier'])
    max_output_characters = int(evaluation_data_configuration['max_ground_truth_characters'] * llm_data_assumptions['input_output_multiplier'])
    
    applicable_llms = {}
    for llm, configuration in llm_configurations.items():
        if configuration['data_unit'] == 'token':
            if zero_shot_tokens > configuration['data_limit']:
                pass
            tokens_left_for_few_shots = configuration['data_limit'] - zero_shot_tokens
            # Assumed up to 10 few shot examples
            max_few_shots = min(10, tokens_left_for_few_shots // max_few_shot_tokens)
            applicable_llms[llm] = {
                'max_few_shots': max_few_shots,
                'max_few_shot_tokens': max_few_shot_tokens,
                'max_output_length': max_output_tokens
            }
        elif configuration['data_unit'] == 'character':
            if zero_shot_characters > configuration['data_limit']:
                pass
            characters_left_for_few_shots = configuration['data_limit'] - zero_shot_characters
            # Assumed up to 10 few shot examples
            max_few_shots = min(10, characters_left_for_few_shots // max_few_shot_characters)
            applicable_llms[llm] = {
                'max_few_shots': max_few_shots,
                'max_few_shot_characters': max_few_shot_characters,
                'max_output_length': max_output_characters
            }
    return applicable_llms

# %%

def estimate_task_creation_cost(evaluation_data_configuration: dict, selected_llms: list, num_prompts: int) -> tuple:
    """
    Returns estimated min and max task creation cost given evaluation data and selected LLMs

    Inputs:
    evaluation_data_configuration: dict -> contains information about evaluation data contents (e.g., tokens, characters)
    selected_llms: list -> selected LLMs to evaluate during task creation
    num_prompts: int -> number of prompts generated with each prompt generation method

    Outputs
    (min_cost, max_cost): tuple -> min and max estimate of task creation cost
    """
    cost = 0

    # Iterate over each of the selected LLMs
    for llm in selected_llms:
        if llm_configurations[llm]['data_unit'] == 'token':
            instruction_length = llm_data_assumptions['instruction_tokens']
            few_shot_length = evaluation_data_configuration['applicable_llms'][llm]['max_few_shot_tokens']
            input_length = evaluation_data_configuration['max_input_tokens']
            output_length = evaluation_data_configuration['max_ground_truth_tokens']
        elif llm_configurations[llm]['data_unit'] == 'character':
            instruction_length = llm_data_assumptions['assumed_instruction_characters']
            few_shot_length = evaluation_data_configuration['applicable_llms'][llm]['max_few_shot_characters']
            input_length = evaluation_data_configuration['max_input_characters']
            output_length = evaluation_data_configuration['max_ground_truth_characters']
        num_few_shots = evaluation_data_configuration['applicable_llms'][llm]['max_few_shots']
        generation_price = llm_configurations['text-davinci-003']['price_per_data_unit']
        inference_price = llm_configurations[llm]['price_per_data_unit']

        print(f'\nCost for llm: {llm}')

        # STAGE 1
        # Prompt generation based on user objective has instruction string and produces multiple prompt candidates
        prompt_generation_user_objective_cost = (instruction_length + instruction_length * num_prompts) * generation_price
        cost += prompt_generation_user_objective_cost
        print(f'prompt_generation_user_objective_cost: {prompt_generation_user_objective_cost:.2f}')
        
        # Prompt generation based on role play pattern uses few shot examples
        prompt_generation_pattern_role_play_cost = (instruction_length + 500 + instruction_length * num_prompts) * generation_price
        cost += prompt_generation_pattern_role_play_cost
        print(f'prompt_generation_pattern_role_play_cost: {prompt_generation_pattern_role_play_cost:.2f}')
        
        # Prompt generation based on user objective with training data uses up to 5 few shot examples
        prompt_generation_user_objective_training_data_cost = (instruction_length + few_shot_length * 5 + \
            instruction_length * num_prompts) * generation_price
        cost += prompt_generation_user_objective_training_data_cost
        print(f'prompt_generation_user_objective_training_data_cost: {prompt_generation_user_objective_training_data_cost: 0.2f}')
        
        # Prompt variant generations generates 3 variant prompts for each original prompt candidate, then checks for overfitting
        prompt_generation_variants_cost = num_prompts * 3 * (instruction_length + 500 + instruction_length * 3) * generation_price
        cost += prompt_generation_variants_cost
        print(f'prompt_generation_variants_cost: {prompt_generation_variants_cost:.2f}')
        
        # Stage 1 evaluation with adaptive filtering
        stage_one_evaluation_cost = (instruction_length + input_length + output_length) * \
            ((30 * 3) + (15 * 3) + (7 * 4)) * inference_price
        cost += stage_one_evaluation_cost
        print(f'stage_one_evaluation_cost: {stage_one_evaluation_cost:.2f}')
        
        # STAGE 2 - Few shots
        # Few shots generation does not cost anything. Evaluation of new few shot prompts incurs fee
        stage_two_evaluation_cost = 5 * (instruction_length + num_few_shots * few_shot_length + input_length + output_length) * 10 * inference_price
        cost += stage_two_evaluation_cost
        print(f'stage_two_evaluation_cost: {stage_two_evaluation_cost:.2f}')

        # STAGE 3 - TIM
        # First, add cost of generating feedback on output for 3 rows
        prompt_generation_targeted_improvement_cost = 3 * (instruction_length * 2 + input_length + output_length * 2 + instruction_length) * generation_price
        # Next, add cost of synthesizing three pieces of feedback
        prompt_generation_targeted_improvement_cost += instruction_length * 3 * generation_price
        # Next, add cost of generating new prompt
        prompt_generation_targeted_improvement_cost += instruction_length * 3 * generation_price
        # Next, add cost of checking for overfitting
        prompt_generation_targeted_improvement_cost += instruction_length * 4 * generation_price
        # TIM is run across 3 prompt candidates
        prompt_generation_targeted_improvement_cost *= 3
        cost += prompt_generation_targeted_improvement_cost
        print(f'prompt_generation_targeted_improvement_cost: {prompt_generation_targeted_improvement_cost:.2f}')

        # Stage 3 evaluation, assuming new TIM prompts are all few shot based
        stage_three_evaluation_cost = 3 * (instruction_length + num_few_shots * few_shot_length + input_length + output_length) * 10 * inference_price
        cost += stage_three_evaluation_cost
        print(f'stage_three_evaluation_cost: {stage_three_evaluation_cost:.2f}')

        # STAGE 4 - LLM eval
        # Only costing llm_eval_precision since that uses davinci while llm_eval_ground_truth_comparison uses turbo
        stage_four_evaluation_cost = 3 * (instruction_length * 2 + input_length + output_length + instruction_length) * 10 * generation_price
        cost += stage_four_evaluation_cost
        print(f'stage_four_evaluation_cost: {stage_four_evaluation_cost:.2f}')

        # STAGE 5 - Temperature variation
        # Assuming 4 different temperature variations of a few shot prompt
        stage_five_evaluation_cost = 4 * (instruction_length + num_few_shots * few_shot_length + input_length + output_length) * 10 * inference_price
        cost += stage_five_evaluation_cost
        print(f'stage_five_evaluation_cost: {stage_five_evaluation_cost:.2f}')

    # Compute cost range assuming +/- 25% of estimated cost
    cost_range_factor = 0.25
    cost_range = (round(cost * (1 - cost_range_factor)), round(cost * (1+ cost_range_factor)))

    print('\notal cost range:')
    return cost_range

# %%

# Configurations for each llm
llm_configurations = {
    'gpt-3.5-turbo': {
        'data_unit': 'token',
        'data_limit': 4096,
        'encoding': 'cl100k_base',
        'price_per_data_unit': 0.002 / 1000
    },
    'text-davinci-003': {
        'data_unit': 'token',
        'data_limit': 4097,
        'encoding': 'p50k_base',
        'price_per_data_unit': 0.02 / 1000
        },
}

# %%

# Assumptions for llm data input / output
llm_data_assumptions = {
    'tokens_per_character': 0.3, # assumes 0.75 words / token and 4.7 characters / word
    'instruction_tokens': 250,
    'instruction_characters': 250 / 0.3,
    'input_output_multiplier': 1.1, # multiplier for input and output data length as buffer
    'buffer_tokens': 250,
    'buffer_characters': 250 / 0.3
}
